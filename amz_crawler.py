import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from datetime import datetime


urls = [
    # 蛋白質
    "https://www.amazon.nl/-/en/Nutrition-Standard-Building-Recovery-Glutamine/dp/B000GIQT3I/ref=sr_1_29",
    "https://www.amazon.de/-/en/Nutrition-Standard-Supplements-Glutamine-packaging/dp/B000GIQT3I/ref=sr_1_4?th=1",
    # 洗衣液
    "https://www.amazon.de/-/en/gp/product/B0C1KC39RY/ref=ewc_pr_img_5?smid=A3JWKAKR8XB7XF&th=1",
    "https://www.amazon.de/-/en/gp/product/B0C1KC6BZ1/ref=ewc_pr_img_6?smid=A3JWKAKR8XB7XF&psc=1",
    "https://www.amazon.nl/-/en/Sensation-Detergent-Colourful-Coloured-Laundry/dp/B098QWY2R2/ref=sr_1_20",
    "https://www.amazon.nl/-/en/Robijn-Radiant-Liquid-Detergent-laundry/dp/B098QZ4NPK/ref=sr_1_12",
    # 濕厠紙
    "https://www.amazon.de/-/en/dp/B0CP9NKPVQ/ref=sr_1_34",
    "https://www.amazon.nl/-/en/Washable-Combination-Refreshes-Intimate-Chamomile/dp/B0CP9NKPVQ/ref=sr_1_12",
    "https://www.amazon.nl/-/en/135508/dp/B0CQ58CPPX/ref=sr_1_1",
    "https://www.amazon.de/-/en/Intimate-Flushable-Combination-Cranberry-Prebiotic/dp/B0CP9MRZ1W/ref=sr_1_47",
    "https://www.amazon.nl/-/en/Cleanic-Junior-damp-cloth-combination/dp/B0CP9MFYWX/ref=sr_1_1",
    "https://www.amazon.nl/-/en/Günstig-pieces-sheets-toilet-Sensitiv/dp/B07G8D1FZ3/ref=sr_1_40",
    "https://www.amazon.de/-/en/Günstig-Sheets-Moist-Toilet-Sensitive/dp/B07G8D1FZ3/ref=sr_1_39",
    "https://www.amazon.nl/-/en/wipes-brand-Amazon-MB3ARFresh-packaging/dp/B07V5VR1F7/ref=sr_1_9",
    # Sebamed
    "https://www.amazon.nl/-/en/Sebamed-soap-free-washpiece-150-g/dp/B00JEX75FC/ref=sr_1_3",
    "https://www.amazon.de/-/en/gp/product/B00JEX75FC/ref=ewc_pr_img_1?smid=A3JWKAKR8XB7XF&psc=1",
    "https://www.amazon.nl/-/en/135508/dp/B0CQ58CPPX/ref=sr_1_1",
    "https://www.amazon.de/-/en/gp/product/B0CQ58CPPX/ref=ewc_pr_img_7?smid=A3JWKAKR8XB7XF&psc=1",
    "https://www.amazon.de/-/en/Sebamed-Anti-Hair-Promotes-Activates-Suitable/dp/B0CQ59LKZT/ref=sr_1_2",
    "https://www.amazon.nl/-/en/Sebamed-anti-hair-promotes-activates-suitable/dp/B0CQ59LKZT/ref=sr_1_7",
    "https://www.amazon.de/Sebamed-Protection-Cream-Waterproof-Microplastics/dp/B07RM22KLX?ref_=ast_sto_dp&th=1&psc=1",
    # 地板清潔
    "https://www.amazon.de/-/en/Proper-Professional-Multi-Purpose-Cleaner-Litres/dp/B0B4SGWMV1/ref=sr_1_51",
    "https://www.amazon.de/-/en/276680-Ajax-All-Purpose-Cleaner-Frischeduft/dp/B00ACWXZTY/ref=sr_1_52",
    "https://www.amazon.nl/-/en/Ajax-All-purpose-cleaner-Citrovers-can/dp/B00CJKLV4I/ref=sr_1_75",
    # 洗碗機 Finish,Sun,Somat..
    # 洗碗精
    "https://www.amazon.de/-/en/DE3070/dp/B0CL9QJ6QV/ref=sr_1_5",
    "https://www.amazon.de/-/en/Ecover-ecological-dishwasher-detergent-lemon/dp/B0013G4G5U/ref=sr_1_6",
    "https://www.amazon.nl/-/en/ECOVER-Dishwashing-liquid-Chamomile-Clementine/dp/B08N59LYF5/ref=sr_1_3",


    
    # 其他
    "https://www.amazon.de/-/en/gp/product/B0CS612572/ref=ewc_pr_img_2?smid=A3JWKAKR8XB7XF&psc=1",
    "https://www.amazon.de/-/en/gp/product/B07PXMT4HX/ref=ewc_pr_img_3?smid=A3JWKAKR8XB7XF&psc=1",
    "https://www.amazon.de/-/en/gp/product/B09DTBY7YW/ref=ewc_pr_img_4?smid=A5RSTJ7MXRC4W&psc=1",
    "https://www.amazon.nl/-/en/Oral-B-Expert-Extra-Fresh-Toothpaste/dp/B09DTBY7YW/ref=sr_1_1",

]

# 存储结果的列表
data = []

# pretend to be a real human
headers = {'User-Agent': (
               'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
               '(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36 OPR/111.0.0.0'
        )
    }
for url in urls:
    response = requests.get(url, headers=headers)

    # 確保請求成功
    if response.status_code != 200:
        raise Exception(f"Failed to load page: {response.status_code}")

    # tell bs4 use html解析
    soup = BeautifulSoup(response.text, "html.parser")

    # 抓取商品名稱
    product_name_element = soup.find('span', id='productTitle')
    if product_name_element:
        product_name = product_name_element.text.strip()
    else:
        product_name = "N/A"
    print(product_name)

    # 抓取商品連接
    print(url)

    # 抓取图片 URL
    img_div = soup.find('div', id='imgTagWrapperId')
    img_url = img_div.find('img')['src'] if img_div and img_div.find('img') else "N/A"
    print(img_url)

    # 抓取價格
    price_symbol_element = soup.find('span', class_='a-price-symbol')
    price_whole_element = soup.find('span', class_='a-price-whole')
    price_fraction_element = soup.find('span', class_='a-price-fraction')
    if price_whole_element:
        price = (
            f"{price_symbol_element.text.strip()}"
            f"{price_whole_element.text.strip()}"
            f"{price_fraction_element.text.strip()}"
        )
    else:
        price = "N/A"
    print(price)

    # 抓取單位價格
    unit_price_element = soup.find('span', class_='a-price a-text-price')
    if unit_price_element:
        unit_price = f"{unit_price_element.find('span', class_='a-offscreen').text.strip()}/kg"
    else:
        unit_price = "N/A"
    print(unit_price)

    # 抓取重量
    # 找到包含 "Unit count" 的行
    unit_count_row = soup.find('tr', class_='po-unit_count')
    if unit_count_row:
        # 获取 "Unit count" 的值
        unit_count_value = unit_count_row.find('span', class_='a-size-base po-break-word')
        unit_count = f"{unit_count_value.text.strip()}"
    else:
        unit_count = "N/A"
    print(unit_count)

    # 抓取運輸費
    delivery_info_element = soup.find('span', {'data-csa-c-type': 'element'}).text.strip()
    if delivery_info_element:
        delivery_info = delivery_info_element
        # 如果字符串包含 'Detail'，去掉 'Detail' 之后的内容
        if 'Detail' in delivery_info:
            delivery_info = delivery_info.split('Detail')[0].strip()
        # 如果字符串包含 'Order'，去掉 'Order' 之后的内容
        elif 'Order' in delivery_info:
            delivery_info = delivery_info.split('Order')[0].strip()
    else:
        delivery_info = "N/A"
    print(delivery_info)

# 把獲得的資料放在同一行
    data.append({
        'Product Name': product_name,
        'URL': url,
        "Image URL:": img_url,
        'Price': price,
        'Unit Price': unit_price,
        'Unit Count': unit_count,
        'Delivery': delivery_info
    })

# 获取当前日期
current_date = datetime.now().strftime('%Y%m%d')

# 创建 DataFrame
df = pd.DataFrame(data)
excel_path = f'amazon_{current_date}.xlsx'

# 使用 pandas 将数据写入 Excel 文件
df.to_excel(excel_path, index=False)

# 读取 Excel 文件
wb = load_workbook(excel_path)
ws = wb.active

# 将 DataFrame 的列名写入 Excel 文件
for col_num, column_title in enumerate(df.columns, 1):
    ws.cell(row=1, column=col_num, value=column_title)

# 将 DataFrame 的行数据写入 Excel 文件
for row_num, row_data in enumerate(df.itertuples(), 2):  # 從row2 開始寫入
    for col_num, cell_value in enumerate(row_data[1:], 1):
        ws.cell(row=row_num, column=col_num, value=cell_value)


print(f"Data has been written to {excel_path}")


# ========================================這裏是調整Excel======================================================

# 遍历 B 列並加入超鏈接
for row in ws.iter_rows(min_col=2, max_col=2, min_row=2, max_row=ws.max_row):
    cell = row[0]  # B 列的单元格
    if cell.value:
        url = cell.value
        # 将单元格内容设置为超链接
        cell.hyperlink = url
        cell.font = Font(color="0000FF", underline="single")  # 设置超链接样式

print("Hyperlinks have been added.")

# 遍历 C 列並加入圖片
for row_num in range(2, ws.max_row + 1):  # C2 開始算，因为第一行是标题
    cell_value = ws.cell(row=row_num, column=3).value  # 获取 C 列的单元格值（图片 URL）
    if cell_value:  # 确保单元格不为空
        try:
            image_response = requests.get(cell_value)
            if image_response.status_code == 200:
                img = Image(BytesIO(image_response.content))
                img.width, img.height = (100, 100)  # 調整圖片大小
                cell_location = f'C{row_num}'
                img.anchor = cell_location
                ws.add_image(img)
        except Exception as e:
            print(f"Failed to load image from {cell_value}: {e}")

    # 设置行的高度
    ws.row_dimensions[row_num].height = 89
# 设置列的宽度
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 15

# Text wrapping A B列
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):  # 在A,B列過一遍
    for cell in row:
        if cell.column_letter in ['A', 'B']:
            cell.alignment = Alignment(wrap_text=True)

print("Images have been added.")
def auto_adjust_column_width(ws, start_col):
    """
    Adjust the widths of columns starting from a specified column based on the content.
    """
    for col in ws.iter_cols(min_col=start_col, max_col=ws.max_column):
        max_length = 0
        column_letter = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                # Check the length of the cell value and update max_length
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except AttributeError:
                pass
        # Add some padding to the column width
        adjusted_width = max_length  # + 2
        ws.column_dimensions[column_letter].width = adjusted_width

# 對齊D以後的内容
auto_adjust_column_width(ws, start_col=4)  # Column D is the 4th column

wb.save(excel_path)

print("Excel has been adjusted.")
